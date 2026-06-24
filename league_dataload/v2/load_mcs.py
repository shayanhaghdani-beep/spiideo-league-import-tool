"""Load a filled Main Camera Sheet (MCS) into structured club records.

The MCS has a ~10-row instruction preamble, then a header row containing
``Team Name``, then one row per club. Each club row carries a subscription, up
to 7 camera blocks (Scene / Type / Position of Field / Installation-Calibration
Status), a contact, and a billing address (== shipping per the sheet's note).

Stdlib only. Works on a CSV export of the sheet; the Google Sheets reader
(Phase A/B live) feeds the same row lists into ``parse_rows``.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

MAX_CAMERAS = 7
HEADER_MARKER = "Team Name"


@dataclass
class Camera:
    index: int
    scene: str = ""
    type: str = ""
    position: str = ""
    install_status: str = ""

    @property
    def filled(self) -> bool:
        return bool(self.type.strip() or self.scene.strip())


@dataclass
class ClubRecord:
    team_name: str = ""
    order_name: str = ""
    gender: str = ""
    sport: str = ""
    arena: str = ""
    customer_sf_id: str = ""
    opportunity_sf_id: str = ""
    effective_start: str = ""
    contact_name: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    billing_country: str = ""
    billing_street: str = ""
    billing_postal: str = ""
    billing_state: str = ""
    billing_city: str = ""
    tax_id: str = ""
    subscription: str = ""
    cameras: list[Camera] = field(default_factory=list)

    @property
    def email_domain(self) -> str:
        e = self.contact_email.strip()
        return e.split("@")[-1].lower() if "@" in e else ""

    @property
    def active_cameras(self) -> list[Camera]:
        return [c for c in self.cameras if c.filled]


def _find_header_row(rows: list[list[str]]) -> int:
    for i, r in enumerate(rows):
        if any((c or "").strip() == HEADER_MARKER for c in r):
            return i
    raise ValueError(f"No header row containing {HEADER_MARKER!r} found")


def _col_index(header: list[str]) -> dict[str, int]:
    """Map a normalised header label -> column index (first occurrence)."""
    idx: dict[str, int] = {}
    for i, raw in enumerate(header):
        key = " ".join((raw or "").split()).strip()
        if key and key not in idx:
            idx[key] = i
    return idx


def _detect_camera_count(idx: dict[str, int]) -> int:
    """Highest 'Camera N' index present in the header. Handles sheets with more
    than the default 7 slots (e.g. an 8th 'Spiideo Stream Encoder' column).
    Falls back to MAX_CAMERAS if none are found."""
    mx = 0
    for key in idx:
        m = re.match(r"Camera (\d+)\b", key)
        if m:
            mx = max(mx, int(m.group(1)))
    return mx or MAX_CAMERAS


def _get(row: list[str], idx: dict[str, int], label: str) -> str:
    i = idx.get(label)
    if i is None or i >= len(row):
        return ""
    return (row[i] or "").strip()


def _parse_camera(row: list[str], idx: dict[str, int], n: int) -> Camera:
    # The WHL header uses a trailing space on some labels ("Camera 3 Scene ");
    # _col_index already collapses internal whitespace, so look up both forms.
    def grab(suffix: str) -> str:
        for label in (f"Camera {n} {suffix}", f"Camera {n} {suffix} "):
            v = _get(row, idx, " ".join(label.split()))
            if v:
                return v
        return ""
    return Camera(
        index=n,
        scene=grab("Scene"),
        type=grab("Type"),
        position=grab("Position of Field"),
        install_status=grab("Installation/Calibration Status"),
    )


def parse_rows(rows: list[list[str]]) -> list[ClubRecord]:
    """Parse raw sheet rows (incl. preamble) into ClubRecords."""
    h = _find_header_row(rows)
    idx = _col_index(rows[h])
    n_cams = _detect_camera_count(idx)
    records: list[ClubRecord] = []
    for row in rows[h + 1:]:
        team = _get(row, idx, "Team Name")
        if not team:
            continue
        rec = ClubRecord(
            team_name=team,
            order_name=_get(row, idx, "Order Name"),
            gender=_get(row, idx, "Team Gender"),
            sport=_get(row, idx, "Sport"),
            arena=_get(row, idx, "Arena Name"),
            customer_sf_id=_get(row, idx, "Customer SF ID"),
            opportunity_sf_id=_get(row, idx, "Opportunity SF ID"),
            effective_start=_get(row, idx, "Effective Start Date"),
            contact_name=_get(row, idx, "Contact Name"),
            contact_email=_get(row, idx, "Contact Email"),
            contact_phone=_get(row, idx, "Contact Number"),
            billing_country=_get(row, idx, "Billing Country"),
            billing_street=_get(row, idx, "Billing Street"),
            billing_postal=_get(row, idx, "Billing Postal Code/Zip Code"),
            billing_state=_get(row, idx, "Billing Province/State"),
            billing_city=_get(row, idx, "Billing City"),
            tax_id=_get(row, idx, "Tax ID"),
            subscription=_get(row, idx, "Spiideo Subscription"),
            cameras=[_parse_camera(row, idx, n) for n in range(1, n_cams + 1)],
        )
        records.append(rec)
    return records


def load_mcs(path: str | Path) -> list[ClubRecord]:
    """Load from CSV or .xlsx (auto-detected by extension)."""
    if str(path).lower().endswith(".xlsx"):
        return parse_rows(_xlsx_rows(path, "Main Camera Sheet"))
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    return parse_rows(rows)


def _xlsx_rows(path: str | Path, sheet: str) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    out: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if c is None else str(c) for c in row])
    return out


def load_config(path: str | Path) -> dict[str, str]:
    """Read the Config tab (label in col A -> value in col B) from a filled xlsx."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    if "Config" not in wb.sheetnames:
        return {}
    cfg = wb["Config"]
    out: dict[str, str] = {}
    for row in cfg.iter_rows(min_col=1, max_col=2, values_only=True):
        label, value = row[0], (row[1] if len(row) > 1 else None)
        if label:
            out[str(label).strip()] = "" if value is None else str(value).strip()
    return out
