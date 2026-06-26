"""Phase A: generate the fillable Main Camera Sheet as a local .xlsx.

Native Excel dropdowns (data validation) for subscription, camera type, position,
sport, gender, and statuses. Green input columns, an instruction preamble, a
product-list hyperlink, and a Config tab carrying the Step-1 choices so Phase B
needs no re-prompting. No Google creds needed -- the operator uploads to Drive.

Dropdown sources: product names from the pricebook; picklist values pulled from
SF (kept here as constants so the generator runs offline).

Requires: openpyxl (Phase-A-only dependency).
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAX_CAMERAS = 7
PRICE_LIST_URL = ("https://docs.google.com/spreadsheets/d/"
                  "1-Vds92avq9cD-CT82i-WDUOGwJgaTSTxg-VYxwe_w5Y/edit")

# Pulled live from SF (active picklist values).
POSITIONS = [
    "Center", "Farcenter", "Left behind goal", "Right behind goal", "Left ceiling",
    "Right ceiling", "Left goal line", "Right goal line", "High Home", "First Base",
    "Third Base", "Center Field", "Left Blue Line", "Right Blue Line", "Scoreboard Cam",
    "Left Corner Cam", "Right Corner Cam", "Net", "Near Sideline", "Far Sideline",
    "Left Baseline", "Right Baseline", "Center Ice", "Left Goal", "Right Goal",
    "High Center", "Half Court", "Behind basket", "Above basket", "Other",
]
SPORTS = [
    "Football/Soccer", "Ice Hockey", "Basketball", "Baseball", "Softball", "Lacrosse",
    "Handball", "Field Hockey", "Futsal", "American Football", "Volleyball", "Cricket",
    "Rugby", "Rugby League", "Floorball", "Bandy", "Multi-Sport", "Water Polo",
    "Netball", "Tennis", "Other",
]
GENDERS = ["Mens", "Womens", "Mens and Womens"]
INSTALL_STATUS = ["Not started", "Scheduled", "Installed", "Calibrated"]
SHIPMENT_STATUS = ["Not shipped", "Shipped", "Delivered"]
CURRENCIES = ["EUR", "USD", "GBP", "SEK"]
PRICING_MODES = ["list", "discount", "free"]
# Subscription billing/price period — Annual by default unless the operator picks otherwise.
BILLING_PERIODS = ["Annual", "Monthly", "Quarterly", "Biannual", "End of term"]
STRUCTURES = ["1 - cameras + subscriptions", "2 - cameras only",
              "3 - cameras to league (no club opps)"]

GREEN = PatternFill("solid", fgColor="C6EFCE")
HDR = PatternFill("solid", fgColor="1F4E2C")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=12)

# (column header, gets dropdown name or None, is rep-input/green)
# Order-generation columns ONLY. Installation/shipping status tracking lives in
# the CSM's separate consolidated installation tracker (decoupled 2026-06-16,
# Nate/Egil/Julia): Tracking Link, Shipment Status, and per-camera
# Installation/Calibration Status are intentionally NOT part of the order sheet.
def _columns() -> list[tuple[str, str | None, bool]]:
    cols: list[tuple[str, str | None, bool]] = [
        ("Team Name", None, True), ("Order Name", None, True),
        ("Team Gender", "gender", True), ("Sport", "sport", True),
        ("Number Of Products", None, True), ("Arena Name", None, True),
        ("Customer SF ID", None, True), ("Opportunity SF ID", None, True),
        ("Effective Start Date", None, True), ("Contact Name", None, True),
        ("Contact Email", None, True), ("Contact Number", None, True),
        ("Billing Country", None, True), ("Billing Street", None, True),
        ("Billing Postal Code/Zip Code", None, True), ("Billing Province/State", None, True),
        ("Billing City", None, True), ("Tax ID", None, True),
        ("Spiideo Subscription", "sub", True),
        # "yes" -> the opp's Order Notes gets "add to SvFF LE" (Shayan, 2026-06-24).
        ("SvFF League Exchange", "yesno", True),
    ]
    for n in range(1, MAX_CAMERAS + 1):
        cols += [
            (f"Camera {n} Scene", None, True),
            (f"Camera {n} Type", "camera", True),
            (f"Camera {n} Position of Field", "position", True),
        ]
    return cols


def _load_products(pricebook_csv: str | Path) -> tuple[list[str], list[str]]:
    """Return (subscriptions, cameras) -- unique product names by family."""
    subs, cams = set(), set()
    with open(pricebook_csv, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            fam = (row.get("Family") or "").lower()
            name = (row.get("Product: Product Name") or "").strip()
            if not name:
                continue
            if "subscription" in fam:
                subs.add(name)
            elif "camera" in fam or "purchase" in fam:
                cams.add(name)
    return sorted(subs), sorted(cams)


def _write_list(ws, col_letter: str, header: str, values: list[str]) -> str:
    ws[f"{col_letter}1"] = header
    for i, v in enumerate(values, start=2):
        ws[f"{col_letter}{i}"] = v
    last = len(values) + 1
    return f"Lists!${col_letter}$2:${col_letter}${last}"


def generate(pricebook_csv: str | Path, out_path: str | Path, *,
             league_name: str = "", rows: int = 60) -> Path:
    subs, cams = _load_products(pricebook_csv)
    wb = Workbook()

    # ---- hidden Lists sheet (dropdown sources) ----
    lists = wb.create_sheet("Lists")
    ranges = {
        "sub": _write_list(lists, "A", "Subscriptions", subs),
        "camera": _write_list(lists, "B", "Cameras", cams),
        "position": _write_list(lists, "C", "Positions", POSITIONS),
        "sport": _write_list(lists, "D", "Sports", SPORTS),
        "gender": _write_list(lists, "E", "Genders", GENDERS),
        "yesno": _write_list(lists, "F", "YesNo", ["yes", "no"]),
    }
    lists.sheet_state = "hidden"

    # ---- main sheet ----
    ws = wb.active
    ws.title = "Main Camera Sheet"
    title = f"Spiideo League Import — Main Camera Sheet" + (f" — {league_name}" if league_name else "")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Fill the GREEN columns, one row per team/club. Billing address = shipping address."
    ws["A3"] = "Product List (reference):"
    ws["B3"].hyperlink = PRICE_LIST_URL
    ws["B3"].value = "Open Price List 2026"
    ws["B3"].font = Font(color="0563C1", underline="single")
    header_row = 5

    cols = _columns()
    for c, (name, dd, green) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = HDR
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = max(14, min(28, len(name) + 2))
        # green input fill for the data rows
        if green:
            for r in range(header_row + 1, header_row + 1 + rows):
                ws.cell(row=r, column=c).fill = GREEN
        # dropdown
        if dd:
            dv = DataValidation(type="list", formula1=ranges[dd], allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letter}{header_row+1}:{letter}{header_row+rows}")
    ws.freeze_panes = f"A{header_row+1}"

    # ---- config tab (Step-1 choices, read by Phase B) ----
    cfg = wb.create_sheet("Config")
    cfg["A1"] = "CONFIG — set before sending to sales (drives the SF import)"
    cfg["A1"].font = TITLE_FONT
    entries = [
        ("Deal Structure", STRUCTURES, "1 - cameras + subscriptions"),
        ("Master Opportunity ID", None, ""),
        ("League", None, league_name),   # opp name = "{Team} - {League} Cam Order"
        ("Currency", CURRENCIES, "EUR"),
        ("Team Gender (competition)", GENDERS, "Mens"),
        ("Subscription pricing", PRICING_MODES, "list"),
        ("Cameras pricing", PRICING_MODES, "list"),
        # Subscriptions bill Annually by default; change only if a deal is billed otherwise.
        ("Subscription billing period", BILLING_PERIODS, "Annual"),
        ("Subscription discount price (if discounted)", None, ""),
        ("Voucher setup (league pre-paid the sub?)", ["no", "yes"], "no"),
        # Owner structure (Shayan, 2026-06-19): child opps -> this CSM (full name or
        # User Id); contacts are always owned by the master opp owner. Blank = opps
        # inherit the master opp owner too.
        ("Opportunity Owner (CSM full name or User Id)", None, ""),
        # OPT-IN, default no (Shayan, 2026-06-22): "yes" adds per camera×region shipping
        # lines to the MASTER opp ($0 sales, cost from product). Sales reps usually handle
        # shipping themselves -> leave "no" unless asked.
        ("Master-opp shipping lines (per camera × region)?", ["no", "yes"], "no"),
    ]
    for i, (label, opts, default) in enumerate(entries, start=3):
        cfg[f"A{i}"] = label
        cfg[f"A{i}"].font = Font(bold=True)
        cfg[f"B{i}"] = default
        cfg[f"B{i}"].fill = GREEN
        if opts:
            dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"', allow_blank=True)
            cfg.add_data_validation(dv)
            dv.add(f"B{i}")
    cfg["A15"] = "Camera discount prices (per model, if cameras = discounted):"
    cfg["A15"].font = Font(bold=True)
    cfg["A16"] = "Camera model"
    cfg["B16"] = "Unit price"
    for j, cam in enumerate(cams[:0], start=17):  # left blank; operator adds models used
        cfg[f"A{j}"] = cam
    cfg.column_dimensions["A"].width = 42
    cfg.column_dimensions["B"].width = 28

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return Path(out_path)
